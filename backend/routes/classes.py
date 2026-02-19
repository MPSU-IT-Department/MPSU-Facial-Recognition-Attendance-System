from flask import Blueprint, render_template, redirect, url_for, request, jsonify, flash, current_app
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from utils.timezone import get_pst_now, pst_now_naive
import calendar
import json
import os
import uuid
from werkzeug.utils import secure_filename
from sqlalchemy import or_, func
from models import User, Class, Student, Enrollment, AttendanceRecord, InstructorAttendance, AttendanceLog, FaceEncoding, ClassSession, Course, SystemSettings
from extensions import db
from forms import ClassForm, EnrollmentForm
from decorators import admin_required
from exceptions import AttendanceValidationError
from sqlalchemy.exc import IntegrityError
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
classes_bp = Blueprint('classes', __name__, url_prefix='/classes')

def _get_payload_value(payload, *keys, default=None):
    if not payload:
        return default
    for key in keys:
        if key in payload and payload[key] not in (None, ''):
            return payload[key]
    return default

def _parse_optional_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None

def _parse_optional_time(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.time()
    if hasattr(value, 'hour') and hasattr(value, 'minute'):
        return value
    text = str(value).strip()
    for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p'):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None

def _serialize_class_payload(cls, instructor_name='Unassigned', substitute_name=None, enrolled_count=0, course_name='Unknown'):
    class_name = cls.class_name or cls.description
    date_value = cls.class_date.isoformat() if cls.class_date else None
    time_value = cls.class_time.strftime('%H:%M:%S') if cls.class_time else None
    return {
        'id': cls.id,
        'classId': cls.id,
        'ClassID': cls.id,
        'classCode': cls.class_code,
        'ClassCode': cls.class_code,
        'className': class_name,
        'ClassName': class_name,
        'description': cls.description,
        'Date': date_value,
        'date': date_value,
        'Time': time_value,
        'time': time_value,
        'roomNumber': cls.room_number,
        'RoomNumber': cls.room_number,
        'schedule': cls.schedule,
        'term': cls.term,
        'Term': cls.term,
        'schoolYear': cls.school_year,
        'SchoolYear': cls.school_year,
        'courseId': cls.course_id,
        'CourseID': cls.course_id,
        'courseName': course_name,
        'instructorId': cls.instructor_id,
        'InstructorID': cls.instructor_id,
        'instructorName': instructor_name,
        'substituteInstructorId': cls.substitute_instructor_id,
        'substituteInstructorName': substitute_name,
        'enrolledCount': enrolled_count,
    }

def parse_schedule_string(schedule_string):
    """Parses a schedule string like 'MTW 10:00 AM-12:00 PM, F 2:00 PM-3:00 PM'
       into a list of dictionaries: [{'days': ['M', 'T', 'W'], 'start': '10:00', 'end': '12:00'}, ...]
       Returns an empty list if parsing fails or input is invalid.
    """
    time_slots = []
    if not schedule_string:
        return time_slots
    slots_str = schedule_string.split(',')
    for slot_str in slots_str:
        try:
            match = re.match('^([MWFSuTh]+)\\s+(\\d{1,2}:\\d{2})\\s*(AM|PM)\\s*-\\s*(\\d{1,2}:\\d{2})\\s*(AM|PM)$', slot_str.strip())
            if not match:
                continue
            days_str, start_time_12_str, start_ampm, end_time_12_str, end_ampm = match.groups()
            days = []
            i = 0
            while i < len(days_str):
                if i < len(days_str) - 1:
                    two_char = days_str[i:i + 2]
                    if two_char == 'Th':
                        days.append(two_char)
                        i += 2
                        continue
                    elif two_char == 'Su':
                        days.append(two_char)
                        i += 2
                        continue
                single_char = days_str[i]
                if single_char in ['M', 'T', 'W', 'F', 'S']:
                    days.append(single_char)
                    i += 1
                else:
                    raise ValueError(f'Invalid day character: {single_char}')

            def convert_to_24hr(time_str, ampm):
                hour, minute = map(int, time_str.split(':'))
                if ampm == 'PM' and hour != 12:
                    hour += 12
                elif ampm == 'AM' and hour == 12:
                    hour = 0
                return f'{hour:02d}:{minute:02d}'
            start_time_24hr = convert_to_24hr(start_time_12_str, start_ampm)
            end_time_24hr = convert_to_24hr(end_time_12_str, end_ampm)
            start_dt = datetime.strptime(start_time_24hr, '%H:%M')
            end_dt = datetime.strptime(end_time_24hr, '%H:%M')
            if end_dt <= start_dt:
                end_dt = end_dt + timedelta(days=1)
            duration = end_dt - start_dt
            if duration.total_seconds() > 24 * 3600:
                raise ValueError(f'Class duration cannot exceed 24 hours')
            if start_time_24hr and end_time_24hr:
                time_slots.append({'days': days, 'start': start_time_24hr, 'end': end_time_24hr, 'is_overnight': end_dt > start_dt + timedelta(days=1)})
        except Exception as e:
            continue
    return time_slots

def check_schedule_conflict(room_number, schedule_string, existing_classes, class_id_to_exclude=None):
    """Checks if the given schedule conflicts with existing classes in the same room.

       Args:
           room_number (str): The room number to check.
           schedule_string (str): The schedule string for the new/updated class.
           existing_classes (list): A list of Class objects to check against.
           class_id_to_exclude (int, optional): The ID of the class being updated,
                                             to exclude it from the conflict check.

       Returns:
           tuple: (bool, str) - True if conflict exists with a message, False otherwise.
    """
    return (False, 'Schedule conflict checks are disabled.')
    new_schedule_slots = parse_schedule_string(schedule_string)
    if not new_schedule_slots:
        return (True, "Invalid schedule format provided. Please use the format: 'MTW 10:00 AM-12:00 PM, F 2:00 PM-3:00 PM'")
    for existing_class in existing_classes:
        if class_id_to_exclude is not None and existing_class.id == class_id_to_exclude:
            continue
        if existing_class.room_number == room_number and existing_class.schedule == schedule_string:
            return (True, f'Duplicate schedule found in Room {room_number}. This exact schedule already exists for class {existing_class.class_code}.')
    conflicts = []
    for existing_class in existing_classes:
        if class_id_to_exclude is not None and existing_class.id == class_id_to_exclude:
            continue
        if existing_class.room_number == room_number:
            existing_schedule_slots = parse_schedule_string(existing_class.schedule)
            if not existing_schedule_slots:
                continue
            for new_slot in new_schedule_slots:
                for existing_slot in existing_schedule_slots:
                    common_days = set(new_slot['days']).intersection(existing_slot['days'])
                    if common_days:
                        new_start = datetime.strptime(new_slot['start'], '%H:%M')
                        new_end = datetime.strptime(new_slot['end'], '%H:%M')
                        existing_start = datetime.strptime(existing_slot['start'], '%H:%M')
                        existing_end = datetime.strptime(existing_slot['end'], '%H:%M')
                        if new_start < existing_end and existing_start < new_end:
                            days_overlap_str = ', '.join(sorted(list(common_days)))
                            conflict_msg = f"Schedule conflict in Room {room_number}:\n- Days: {days_overlap_str}\n- Time: {new_slot['start']}-{new_slot['end']}\n- Conflicts with class {existing_class.class_code} ({existing_slot['start']}-{existing_slot['end']})"
                            conflicts.append(conflict_msg)
    if conflicts:
        return (True, 'Multiple conflicts found:\n' + '\n'.join(conflicts))
    return (False, 'No schedule conflicts detected.')

def check_instructor_schedule_conflict(instructor_id, schedule_string, existing_classes, class_id_to_exclude=None):
    """Checks if the given schedule conflicts with existing classes for the same instructor.

       Args:
           instructor_id (int): The ID of the instructor to check.
           schedule_string (str): The schedule string for the new/updated class.
           existing_classes (list): A list of Class objects to check against.
           class_id_to_exclude (int, optional): The ID of the class being updated,
                                             to exclude it from the conflict check.

       Returns:
           tuple: (bool, str) - True if conflict exists with a message, False otherwise.
    """
    try:
        instructor_id = int(instructor_id)
    except (TypeError, ValueError):
        return (True, 'Invalid instructor selection for conflict check.')
    new_schedule_slots = parse_schedule_string(schedule_string)
    if not new_schedule_slots:
        return (True, "Invalid schedule format provided. Please use the format: 'MTW 10:00 AM-12:00 PM, F 2:00 PM-3:00 PM'")
    conflicts = []
    for existing_class in existing_classes:
        if class_id_to_exclude is not None and existing_class.id == class_id_to_exclude:
            continue
        if existing_class.instructor_id == instructor_id:
            existing_schedule_slots = parse_schedule_string(existing_class.schedule)
            if not existing_schedule_slots:
                continue
            for new_slot in new_schedule_slots:
                for existing_slot in existing_schedule_slots:
                    common_days = set(new_slot['days']).intersection(existing_slot['days'])
                    if common_days:
                        new_start = datetime.strptime(new_slot['start'], '%H:%M')
                        new_end = datetime.strptime(new_slot['end'], '%H:%M')
                        existing_start = datetime.strptime(existing_slot['start'], '%H:%M')
                        existing_end = datetime.strptime(existing_slot['end'], '%H:%M')
                        if new_start < existing_end and existing_start < new_end:
                            days_overlap_str = ', '.join(sorted(list(common_days)))
                            conflict_msg = f"Instructor schedule conflict:\n- Days: {days_overlap_str}\n- Time: {new_slot['start']}-{new_slot['end']}\n- Conflicts with class {existing_class.class_code} ({existing_slot['start']}-{existing_slot['end']})"
                            conflicts.append(conflict_msg)
    if conflicts:
        return (True, 'Multiple instructor conflicts found:\n' + '\n'.join(conflicts))
    return (False, 'No instructor schedule conflicts detected.')

def standardize_schedule_days(schedule_string):
    """Standardizes the order of days in a schedule string.
       Example: 'TMW 10:00 AM-12:00 PM' becomes 'MTW 10:00 AM-12:00 PM'
    """
    if not schedule_string:
        return schedule_string
    slots = schedule_string.split(',')
    standardized_slots = []
    for slot in slots:
        try:
            parts = slot.strip().split(' ', 1)
            if len(parts) != 2:
                continue
            days_str, time_part = parts
            day_order = {'M': 0, 'T': 1, 'W': 2, 'Th': 3, 'F': 4, 'S': 5, 'Su': 6}
            days = []
            i = 0
            while i < len(days_str):
                if i + 1 < len(days_str) and days_str[i:i + 2] == 'Th':
                    days.append('Th')
                    i += 2
                elif i + 1 < len(days_str) and days_str[i:i + 2] == 'Su':
                    days.append('Su')
                    i += 2
                else:
                    days.append(days_str[i])
                    i += 1
            days.sort(key=lambda x: day_order.get(x, 999))
            standardized_slots.append(f"{''.join(days)} {time_part}")
        except Exception as e:
            standardized_slots.append(slot)
    return ', '.join(standardized_slots)

def validate_schedule_format(schedule_string):
    """Validates the format of a schedule string.
    
    Args:
        schedule_string (str): The schedule string to validate
        
    Returns:
        tuple: (bool, str) - (is_valid, error_message)
    """
    if not schedule_string:
        return (False, 'Schedule cannot be empty')
    slots = schedule_string.split(',')
    if not slots:
        return (False, 'Invalid schedule format')
    for slot in slots:
        slot = slot.strip()
        if not re.match('^[MWFSuTh]+\\s+\\d{1,2}:\\d{2}\\s*(AM|PM)\\s*-\\s*\\d{1,2}:\\d{2}\\s*(AM|PM)$', slot):
            return (False, f"Invalid schedule format in slot: {slot}\nExpected format: DAYS TIME-TIME (e.g., 'MTW 10:00 AM-12:00 PM' or 'TTh 2:30 PM-4:30 PM')")
        days_part = slot.split()[0]
        valid_days = {'M', 'T', 'W', 'Th', 'F', 'S', 'Su'}
        parsed_days = []
        i = 0
        while i < len(days_part):
            if i < len(days_part) - 1:
                two_char = days_part[i:i + 2]
                if two_char == 'Th':
                    parsed_days.append(two_char)
                    i += 2
                    continue
                elif two_char == 'Su':
                    parsed_days.append(two_char)
                    i += 2
                    continue
            single_char = days_part[i]
            if single_char in ['M', 'T', 'W', 'F', 'S']:
                parsed_days.append(single_char)
                i += 1
            else:
                return (False, f"Invalid day '{single_char}' in schedule slot: {slot}\nValid days are: M, T, W, Th, F, S, Su")
        time_parts = slot.split()[1:]
        if len(time_parts) < 3:
            return (False, f"Invalid time format in slot: {slot}\nExpected format: TIME-TIME (e.g., '10:00 AM-12:00 PM' or '11:45 PM-1:00 AM')")
        try:
            time_string = ' '.join(time_parts)
            if '-' not in time_string:
                return (False, f'Invalid time format in slot: {slot}\nMissing dash between start and end times')
            start_time_str, end_time_str = time_string.split('-', 1)
            start_time_str = start_time_str.strip()
            end_time_str = end_time_str.strip()

            def convert_to_24hr(time_str):
                time_obj = datetime.strptime(time_str, '%I:%M %p')
                return time_obj.strftime('%H:%M')
            start_24hr = convert_to_24hr(start_time_str)
            end_24hr = convert_to_24hr(end_time_str)
            start_dt = datetime.strptime(start_24hr, '%H:%M')
            end_dt = datetime.strptime(end_24hr, '%H:%M')
            if end_dt <= start_dt:
                end_dt = end_dt + timedelta(days=1)
            duration = end_dt - start_dt
            if duration.total_seconds() > 24 * 3600:
                return (False, f'Class duration cannot exceed 24 hours in slot: {slot}')
        except ValueError as e:
            return (False, f'Invalid time format in slot: {slot}\nError: {str(e)}')
    return (True, '')

def parse_instructor_identifier(raw_value, label='instructor'):
    """Normalize instructor IDs coming from JSON payloads."""
    if raw_value is None or raw_value == '' or str(raw_value).lower() == 'null':
        return None
    try:
        return int(raw_value)
    except (TypeError, ValueError):
        raise ValueError(f'Invalid {label} selection.')

@classes_bp.route('/schedule', methods=['GET'])
@login_required
def schedule():
    form = ClassForm()
    instructors = User.query.filter_by(role='instructor').all()
    form.instructor_id.choices = [(i.id, f'{i.first_name} {i.last_name}') for i in instructors]
    return render_template('admin/classes.html', form=form)

@classes_bp.route('/debug-info', methods=['GET'])
@login_required
def debug_info():
    """Debug endpoint to return information about the system state."""
    try:
        from flask import session as flask_session
        import sys
        import platform
        import flask
        db_uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', 'Not configured')
        db_uri_safe = db_uri.split('@')[0] + '@' + db_uri.split('@')[1].split('/')[0] + '/****' if '@' in db_uri else db_uri
        user_info = {'id': current_user.id if current_user else None, 'username': current_user.username if current_user else None, 'role': current_user.role if current_user else None, 'authenticated': current_user.is_authenticated if current_user else False}
        class_count = Class.query.count()
        user_count = User.query.count()
        student_count = Student.query.count()
        enrollment_count = Enrollment.query.count()
        debug_info = {'timestamp': get_pst_now().isoformat(), 'python_version': sys.version, 'platform': platform.platform(), 'flask_version': flask.__version__, 'session_keys': list(flask_session.keys()), 'current_user': user_info, 'database': {'uri': db_uri_safe, 'class_count': class_count, 'user_count': user_count, 'student_count': student_count, 'enrollment_count': enrollment_count}}
        return jsonify(debug_info)
    except Exception as e:
        return (jsonify({'error': str(e)}), 500)

@classes_bp.route('/api/list', methods=['GET'])
def get_classes():
    api_key = request.headers.get('X-API-Key')
    if api_key and api_key == current_app.config.get('API_KEY'):
        classes = Class.query.all()
    elif current_user.is_authenticated:
        if current_user.role == 'instructor':
            classes = Class.query.filter_by(instructor_id=current_user.id).all()
        else:
            classes = Class.query.all()
    else:
        return (jsonify({'success': False, 'message': 'Unauthorized'}), 401)
    try:
        class_list = []
        for cls in classes:
            try:
                if cls.instructor_id:
                    instructor = User.query.get(cls.instructor_id)
                    if instructor:
                        instructor_name = f'{instructor.first_name} {instructor.last_name}'
                    else:
                        instructor_name = 'Unknown'
                else:
                    instructor_name = 'Unassigned'
                substitute_name = None
                if cls.substitute_instructor_id:
                    substitute = User.query.get(cls.substitute_instructor_id)
                    if substitute:
                        substitute_name = f'{substitute.first_name} {substitute.last_name}'
                    else:
                        substitute_name = 'Unknown'
                enrolled_count = Enrollment.query.filter_by(class_id=cls.id).count()
                course = Course.query.get(cls.course_id)
                course_name = course.description if course else 'Unknown'
                class_data = _serialize_class_payload(
                    cls,
                    instructor_name=instructor_name,
                    substitute_name=substitute_name,
                    enrolled_count=enrolled_count,
                    course_name=course_name,
                )
                class_list.append(class_data)
            except Exception as e:
                pass
        return jsonify(class_list)
    except Exception as e:
        import traceback
        return (jsonify({'error': str(e)}), 500)

@classes_bp.route('/api/<int:class_id>', methods=['GET'])
@login_required
def get_class(class_id):
    try:
        cls = Class.query.get(class_id)
        if not cls:
            return (jsonify({'success': False, 'message': 'Class not found'}), 404)
        if current_user.role == 'instructor' and cls.instructor_id != current_user.id:
            return (jsonify({'success': False, 'message': 'You do not have permission to view this class'}), 403)
        if cls.instructor_id:
            instructor = User.query.get(cls.instructor_id)
            instructor_name = f'{instructor.first_name} {instructor.last_name}' if instructor else 'Unknown'
        else:
            instructor_name = 'Unassigned'
        if cls.substitute_instructor_id:
            substitute = User.query.get(cls.substitute_instructor_id)
            substitute_name = f'{substitute.first_name} {substitute.last_name}' if substitute else 'Unknown'
        else:
            substitute_name = None
        course = Course.query.get(cls.course_id)
        course_name = course.description if course else 'Unknown'
        enrolled_count = Enrollment.query.filter_by(class_id=cls.id).count()
        enrollments = Enrollment.query.filter_by(class_id=cls.id).all()
        students = []
        for enrollment in enrollments:
            student = Student.query.get(enrollment.student_id)
            if student:
                face_encoding = FaceEncoding.query.filter_by(student_id=student.id).first()
                profile_image = face_encoding.image_path if face_encoding and face_encoding.image_path else None
                students.append({'id': student.id, 'name': f'{student.first_name} {student.last_name}'})
        response = _serialize_class_payload(
            cls,
            instructor_name=instructor_name,
            substitute_name=substitute_name,
            enrolled_count=enrolled_count,
            course_name=course_name,
        )
        response['students'] = students
        return jsonify(response)
    except Exception as e:
        import traceback
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/create', methods=['POST'])
@login_required
@admin_required
def create_class():
    try:
        data = request.get_json() or {}
        class_code = _get_payload_value(data, 'classCode', 'class_code', 'ClassCode')
        class_name = _get_payload_value(data, 'className', 'class_name', 'ClassName', 'description')
        course_id_raw = _get_payload_value(data, 'courseId', 'course_id', 'CourseID')
        schedule_input = _get_payload_value(data, 'schedule', 'Schedule')
        room_number = _get_payload_value(data, 'roomNumber', 'room_number', 'RoomNumber', default='310')
        class_date = _parse_optional_date(_get_payload_value(data, 'date', 'classDate', 'class_date', 'Date'))
        class_time = _parse_optional_time(_get_payload_value(data, 'time', 'classTime', 'class_time', 'Time'))
        if not class_code:
            return (jsonify({'error': 'Missing required field: classCode'}), 400)
        if not class_name:
            return (jsonify({'error': 'Missing required field: description or className'}), 400)
        if course_id_raw in (None, ''):
            return (jsonify({'error': 'Missing required field: courseId'}), 400)
        try:
            course_id = int(course_id_raw)
        except (TypeError, ValueError):
            return (jsonify({'error': 'Invalid courseId'}), 400)
        try:
            instructor_id = parse_instructor_identifier(_get_payload_value(data, 'instructorId', 'instructor_id', 'InstructorID'))
            substitute_instructor_id = parse_instructor_identifier(_get_payload_value(data, 'substituteInstructorId', 'substitute_instructor_id', 'SubstituteInstructorID'), label='substitute instructor')
        except ValueError as exc:
            return (jsonify({'error': str(exc)}), 400)
        existing_class = Class.query.filter(func.lower(Class.class_code) == func.lower(class_code)).first()
        if existing_class:
            return (jsonify({'error': 'Class code already exists'}), 409)
        schedule = None
        if schedule_input:
            schedule = standardize_schedule_days(str(schedule_input))
            is_valid, error_message = validate_schedule_format(schedule)
            if not is_valid:
                return (jsonify({'error': f'Invalid schedule format: {error_message}'}), 400)
            conflict, message = check_schedule_conflict(room_number, schedule, Class.query.all())
            if conflict:
                return (jsonify({'error': message}), 409)
            if instructor_id:
                instructor_conflict, instructor_message = check_instructor_schedule_conflict(instructor_id, schedule, Class.query.all())
                if instructor_conflict:
                    return (jsonify({'error': instructor_message}), 409)
        from models import SystemSettings
        settings = SystemSettings.query.all()
        settings_dict = {s.key: s.value for s in settings}
        default_term = settings_dict.get('semester', '1st semester')
        default_school_year = settings_dict.get('school_year', '2025-2026')
        new_class = Class(
            class_code=class_code,
            class_name=class_name,
            description=class_name,
            class_date=class_date,
            class_time=class_time,
            room_number=room_number,
            schedule=schedule,
            instructor_id=instructor_id,
            substitute_instructor_id=substitute_instructor_id,
            course_id=course_id,
            term=str(_get_payload_value(data, 'term', 'Term', default=default_term)).lower(),
            school_year=str(_get_payload_value(data, 'schoolYear', 'school_year', 'SchoolYear', default=default_school_year)),
            created_at=pst_now_naive(),
        )
        db.session.add(new_class)
        db.session.commit()
        substitute_name = None
        if new_class.substitute_instructor_id:
            substitute = User.query.get(new_class.substitute_instructor_id)
            substitute_name = f'{substitute.first_name} {substitute.last_name}' if substitute else 'Unknown'
        if new_class.instructor_id:
            primary_instructor = User.query.get(new_class.instructor_id)
            instructor_name = f'{primary_instructor.first_name} {primary_instructor.last_name}' if primary_instructor else 'Unknown'
        else:
            instructor_name = 'Unassigned'
        course = Course.query.get(new_class.course_id)
        class_payload = _serialize_class_payload(
            new_class,
            instructor_name=instructor_name,
            substitute_name=substitute_name,
            enrolled_count=0,
            course_name=course.description if course else 'Unknown',
        )
        return (jsonify({'message': 'Class created successfully', 'class': class_payload}), 201)
    except Exception as e:
        db.session.rollback()
        return (jsonify({'error': str(e)}), 500)

@classes_bp.route('/api/update/<int:class_id>', methods=['PUT'])
@login_required
@admin_required
def update_class(class_id):
    cls = Class.query.get(class_id)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    data = request.get_json() or {}
    class_code = _get_payload_value(data, 'classCode', 'class_code', 'ClassCode')
    class_name = _get_payload_value(data, 'className', 'class_name', 'ClassName', 'description')
    schedule_input = _get_payload_value(data, 'schedule', 'Schedule')
    if not data:
        return (jsonify({'success': False, 'message': 'No input data provided for update'}), 400)
    updated_room_number = _get_payload_value(data, 'roomNumber', 'room_number', 'RoomNumber', default='310')
    updated_schedule = cls.schedule
    if schedule_input is not None:
        if schedule_input:
            updated_schedule = standardize_schedule_days(str(schedule_input))
        else:
            updated_schedule = None
    all_existing_classes = Class.query.all()
    if updated_schedule:
        conflict, message = check_schedule_conflict(updated_room_number, updated_schedule, all_existing_classes, class_id_to_exclude=class_id)
        if conflict:
            return (jsonify({'success': False, 'message': message}), 409)
    try:
        updated_instructor_id = parse_instructor_identifier(_get_payload_value(data, 'instructorId', 'instructor_id', 'InstructorID')) if any((k in data for k in ('instructorId', 'instructor_id', 'InstructorID'))) else cls.instructor_id
        updated_substitute_id = parse_instructor_identifier(_get_payload_value(data, 'substituteInstructorId', 'substitute_instructor_id', 'SubstituteInstructorID'), label='substitute instructor') if any((k in data for k in ('substituteInstructorId', 'substitute_instructor_id', 'SubstituteInstructorID'))) else cls.substitute_instructor_id
    except ValueError as exc:
        return (jsonify({'success': False, 'message': str(exc)}), 400)
    if updated_instructor_id and updated_schedule:
        instructor_conflict, instructor_message = check_instructor_schedule_conflict(updated_instructor_id, updated_schedule, all_existing_classes, class_id_to_exclude=class_id)
        if instructor_conflict:
            return (jsonify({'success': False, 'message': instructor_message}), 409)
    if class_code and class_code != cls.class_code:
        existing_class = Class.query.filter_by(class_code=class_code).first()
        if existing_class and existing_class.id != class_id:
            return (jsonify({'success': False, 'message': 'Class code already exists'}), 400)
        cls.class_code = class_code
    if class_name:
        cls.description = class_name
        cls.class_name = class_name
    date_input = _get_payload_value(data, 'date', 'classDate', 'class_date', 'Date')
    if date_input is not None:
        cls.class_date = _parse_optional_date(date_input)
    time_input = _get_payload_value(data, 'time', 'classTime', 'class_time', 'Time')
    if time_input is not None:
        cls.class_time = _parse_optional_time(time_input)
    cls.room_number = updated_room_number
    if schedule_input is not None:
        cls.schedule = updated_schedule
    if any((k in data for k in ('instructorId', 'instructor_id', 'InstructorID'))):
        cls.instructor_id = updated_instructor_id
    if any((k in data for k in ('substituteInstructorId', 'substitute_instructor_id', 'SubstituteInstructorID'))):
        cls.substitute_instructor_id = updated_substitute_id
    course_id_input = _get_payload_value(data, 'courseId', 'course_id', 'CourseID')
    if course_id_input not in (None, ''):
        try:
            new_course_id = int(course_id_input)
        except (ValueError, TypeError):
            return (jsonify({'success': False, 'message': 'Invalid courseId'}), 400)
        course_exists = Course.query.get(new_course_id)
        if not course_exists:
            return (jsonify({'success': False, 'message': 'Course not found'}), 404)
        cls.course_id = new_course_id
    try:
        db.session.commit()
        instructor = User.query.get(cls.instructor_id) if cls.instructor_id else None
        instructor_name = f'{instructor.first_name} {instructor.last_name}' if instructor else 'Unassigned'
        substitute = User.query.get(cls.substitute_instructor_id) if cls.substitute_instructor_id else None
        if cls.substitute_instructor_id:
            substitute_name = f'{substitute.first_name} {substitute.last_name}' if substitute else 'Unknown'
        else:
            substitute_name = None
        enrolled_count = Enrollment.query.filter_by(class_id=cls.id).count()
        course = Course.query.get(cls.course_id) if cls.course_id else None
        class_payload = _serialize_class_payload(
            cls,
            instructor_name=instructor_name,
            substitute_name=substitute_name,
            enrolled_count=enrolled_count,
            course_name=course.description if course else 'Unknown',
        )
        return jsonify({'success': True, 'message': 'Class updated successfully', 'class': class_payload})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/delete/<int:class_id>', methods=['DELETE'])
@login_required
def delete_class(class_id):
    cls = Class.query.get(class_id)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    try:
        AttendanceRecord.query.filter_by(class_id=class_id).delete(synchronize_session=False)
        class_sessions = ClassSession.query.filter_by(class_id=class_id).all()
        for session in class_sessions:
            AttendanceRecord.query.filter_by(class_session_id=session.id).delete(synchronize_session=False)
        ClassSession.query.filter_by(class_id=class_id).delete()
        Enrollment.query.filter_by(class_id=class_id).delete()
        db.session.delete(cls)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Class deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/delete-all', methods=['DELETE'])
@login_required
@admin_required
def delete_all_classes():
    """Delete all classes and related data (sessions, attendance, enrollments)."""
    try:
        AttendanceRecord.query.delete(synchronize_session=False)
        sessions = ClassSession.query.all()
        for session in sessions:
            AttendanceRecord.query.filter_by(class_session_id=session.id).delete(synchronize_session=False)
        ClassSession.query.delete()
        Enrollment.query.delete()
        Class.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'All classes deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/<int:class_id>/students', methods=['GET'])
@login_required
def get_class_students(class_id):
    cls = Class.query.get(class_id)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    if current_user.role == 'instructor' and cls.instructor_id != current_user.id:
        return (jsonify({'success': False, 'message': 'You do not have permission to view this class'}), 403)
    enrollments = Enrollment.query.filter_by(class_id=class_id).all()
    student_list = []
    for enrollment in enrollments:
        student = Student.query.get(enrollment.student_id)
        if student:
            face_encoding = FaceEncoding.query.filter_by(student_id=student.id).first()
            profile_image = face_encoding.image_path if face_encoding and face_encoding.image_path else None
            student_list.append({'id': student.id, 'studentId': student.id, 'StudentID': student.id, 'firstName': student.first_name, 'lastName': student.last_name, 'yearLevel': student.year_level, 'phone': getattr(student, 'phone', ''), 'email': getattr(student, 'email', '') or '', 'enrollmentId': enrollment.id, 'EnrollmentID': enrollment.id, 'enrollmentDate': enrollment.created_at.strftime('%Y-%m-%d'), 'profileImage': profile_image})
    return jsonify(student_list)

@classes_bp.route('/api/<int:class_id>/enroll', methods=['POST'])
@login_required
def enroll_student(class_id):
    if current_user.role != 'instructor':
        return (jsonify({'success': False, 'message': 'Only instructors can enroll students'}), 403)
    cls = Class.query.get(class_id)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    if cls.instructor_id != current_user.id:
        return (jsonify({'success': False, 'message': 'You can only enroll students in classes you teach'}), 403)
    data = request.get_json() or {}
    student_id = _get_payload_value(data, 'studentId', 'student_id', 'StudentID')
    if not student_id:
        return (jsonify({'success': False, 'message': 'Missing or empty studentId'}), 400)
    student = Student.query.get(student_id)
    if not student:
        return (jsonify({'success': False, 'message': 'Student not found'}), 404)
    existing_enrollment = Enrollment.query.filter_by(class_id=class_id, student_id=student_id).first()
    if existing_enrollment:
        return (jsonify({'success': False, 'message': 'Student already enrolled in this class'}), 400)
    enrollment = Enrollment(student_id=student_id, class_id=class_id, school_year=cls.school_year, term=cls.term, enrolled_date=pst_now_naive())
    try:
        db.session.add(enrollment)
        db.session.commit()
        face_encoding = FaceEncoding.query.filter_by(student_id=student.id).first()
        profile_image = face_encoding.image_path if face_encoding and face_encoding.image_path else None
        return jsonify({'success': True, 'message': 'Student enrolled successfully', 'student': {'id': student.id, 'StudentID': student.id, 'firstName': student.first_name, 'lastName': student.last_name, 'yearLevel': student.year_level, 'phone': getattr(student, 'phone', ''), 'email': getattr(student, 'email', '') or '', 'enrollmentId': enrollment.id, 'EnrollmentID': enrollment.id, 'schoolYear': enrollment.school_year, 'term': enrollment.term, 'enrollmentDate': enrollment.enrolled_date.strftime('%Y-%m-%d'), 'profileImage': profile_image}})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/<int:class_id>/unenroll/<int:enrollment_id>', methods=['DELETE'])
@login_required
def unenroll_student_by_enrollment(class_id, enrollment_id):
    if current_user.role != 'instructor':
        return (jsonify({'success': False, 'message': 'Only instructors can unenroll students'}), 403)
    cls = Class.query.get(class_id)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    if cls.instructor_id != current_user.id:
        return (jsonify({'success': False, 'message': 'You can only unenroll students from classes you teach'}), 403)
    enrollment = Enrollment.query.get(enrollment_id)
    if not enrollment:
        return (jsonify({'success': False, 'message': 'Enrollment record not found'}), 404)
    if enrollment.class_id != class_id:
        return (jsonify({'success': False, 'message': 'Enrollment does not belong to this class'}), 400)
    student = Student.query.get(enrollment.student_id)
    student_info = {'id': student.id, 'firstName': student.first_name, 'lastName': student.last_name}
    try:
        db.session.delete(enrollment)
        db.session.commit()
        return jsonify({'success': True, 'message': f"Student {student_info['firstName']} {student_info['lastName']} unenrolled successfully", 'student': student_info})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/<int:class_id>/unenroll/<string:student_id>', methods=['DELETE'])
@login_required
def unenroll_student_by_id(class_id, student_id):
    if current_user.role != 'instructor':
        return (jsonify({'success': False, 'message': 'Only instructors can unenroll students'}), 403)
    cls = Class.query.get(class_id)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    if cls.instructor_id != current_user.id:
        return (jsonify({'success': False, 'message': 'You can only unenroll students from classes you teach'}), 403)
    enrollment = Enrollment.query.filter_by(class_id=class_id, student_id=student_id).first()
    if not enrollment:
        return (jsonify({'success': False, 'message': 'Student not enrolled in this class'}), 404)
    student = Student.query.get(enrollment.student_id)
    student_info = {'id': student.id, 'firstName': student.first_name, 'lastName': student.last_name}
    try:
        db.session.delete(enrollment)
        db.session.commit()
        return jsonify({'success': True, 'message': f"Student {student_info['firstName']} {student_info['lastName']} unenrolled successfully", 'student': student_info})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/enroll', methods=['POST'])
@login_required
def enroll_student_general():
    """Generic enrollment endpoint that doesn't require class ID in the URL path"""
    if current_user.role not in ['instructor', 'admin']:
        return (jsonify({'success': False, 'message': 'Only instructors and administrators can enroll students'}), 403)
    data = request.get_json() or {}
    student_id = _get_payload_value(data, 'studentId', 'student_id', 'StudentID')
    class_id_value = _get_payload_value(data, 'classId', 'class_id', 'ClassID')
    if not student_id or not class_id_value:
        return (jsonify({'success': False, 'message': 'Missing required enrollment data'}), 400)
    cls = Class.query.get(class_id_value)
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    if current_user.role == 'instructor' and cls.instructor_id != current_user.id:
        return (jsonify({'success': False, 'message': 'You can only enroll students in classes you teach'}), 403)
    student = Student.query.get(student_id)
    if not student:
        return (jsonify({'success': False, 'message': 'Student not found'}), 404)
    existing_enrollment = Enrollment.query.filter_by(class_id=class_id_value, student_id=student_id).first()
    if existing_enrollment:
        return (jsonify({'success': False, 'message': 'Student already enrolled in this class'}), 400)
    enrollment = Enrollment(student_id=student_id, class_id=class_id_value, school_year=cls.school_year, term=cls.term, enrolled_date=pst_now_naive())
    try:
        db.session.add(enrollment)
        db.session.commit()
        face_encoding = FaceEncoding.query.filter_by(student_id=student.id).first()
        profile_image = face_encoding.image_path if face_encoding and face_encoding.image_path else None
        return jsonify({'success': True, 'message': 'Student enrolled successfully', 'student': {'id': student.id, 'StudentID': student.id, 'firstName': student.first_name, 'lastName': student.last_name, 'yearLevel': student.year_level, 'phone': getattr(student, 'phone', ''), 'email': getattr(student, 'email', '') or '', 'enrollmentId': enrollment.id, 'EnrollmentID': enrollment.id, 'schoolYear': enrollment.school_year, 'term': enrollment.term, 'enrollmentDate': enrollment.enrolled_date.strftime('%Y-%m-%d'), 'profileImage': profile_image}})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/unenroll', methods=['DELETE'])
@login_required
def unenroll_student_general():
    """Generic unenrollment endpoint that doesn't require class ID in the URL path"""
    if current_user.role not in ['instructor', 'admin']:
        return (jsonify({'success': False, 'message': 'Only instructors and administrators can unenroll students'}), 403)
    data = request.get_json()
    required_fields = ['studentId', 'classId']
    if not data:
        return (jsonify({'success': False, 'message': 'No input data provided for unenrollment'}), 400)
    for field in required_fields:
        if field not in data or not data[field]:
            return (jsonify({'success': False, 'message': f'Missing or empty required field: {field}'}), 400)
    try:
        class_id = int(data['classId'])
        data['classId'] = class_id
    except (ValueError, TypeError):
        return (jsonify({'success': False, 'message': 'Invalid classId format. Must be an integer.'}), 400)
    cls = Class.query.get(data['classId'])
    if not cls:
        return (jsonify({'success': False, 'message': 'Class not found'}), 404)
    if current_user.role == 'instructor' and cls.instructor_id != current_user.id:
        return (jsonify({'success': False, 'message': 'You can only unenroll students from classes you teach'}), 403)
    enrollment = Enrollment.query.filter_by(class_id=data['classId'], student_id=data['studentId']).first()
    if not enrollment:
        return (jsonify({'success': False, 'message': 'Student not enrolled in this class'}), 404)
    try:
        student = Student.query.get(enrollment.student_id)
        student_info = {'id': student.id, 'firstName': student.first_name, 'lastName': student.last_name}
        db.session.delete(enrollment)
        db.session.commit()
        return jsonify({'success': True, 'message': f"Student {student_info['firstName']} {student_info['lastName']} unenrolled successfully", 'student': student_info})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/next-section/<course_code>', methods=['GET'])
@login_required
def get_next_section(course_code):
    """Get the next available section letter for a course"""
    try:
        existing_classes = Class.query.filter(Class.class_code.like(f'{course_code}-%')).all()
        existing_sections = set()
        for cls in existing_classes:
            parts = cls.class_code.split('-')
            if len(parts) > 1:
                section = parts[-1]
                if len(section) == 1 and section.isalpha():
                    existing_sections.add(section.upper())
        for letter in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            if letter not in existing_sections:
                return jsonify({'success': True, 'section': letter})
        return (jsonify({'success': False, 'message': 'No available section letters (A-Z)'}), 400)
    except Exception as e:
        return (jsonify({'success': False, 'message': str(e)}), 500)

@classes_bp.route('/api/check-conflicts', methods=['POST'])
@login_required
def check_conflicts():
    """API endpoint to check for schedule conflicts before saving."""
    data = request.get_json()
    if not data or 'schedule' not in data:
        return (jsonify({'success': False, 'message': 'Missing required field: schedule'}), 400)
    room_number = (data.get('roomNumber') or '310').strip()
    if not room_number:
        room_number = '310'
    schedule_string = data['schedule'].strip()
    is_valid, error_message = validate_schedule_format(schedule_string)
    if not is_valid:
        return (jsonify({'success': False, 'message': error_message}), 400)
    schedule_string = standardize_schedule_days(schedule_string)
    existing_classes = Class.query.all()
    conflict, message = check_schedule_conflict(room_number, schedule_string, existing_classes)
    instructor_conflict = False
    instructor_message = ''
    if 'instructorId' in data and data['instructorId']:
        instructor_conflict, instructor_message = check_instructor_schedule_conflict(data['instructorId'], schedule_string, existing_classes)
    has_conflict = conflict or instructor_conflict
    combined_message = message
    if instructor_conflict:
        combined_message += '\n' + instructor_message if conflict else instructor_message
    return (jsonify({'success': not has_conflict, 'message': combined_message}), 200 if not has_conflict else 409)

@classes_bp.route('/api/export-classes', methods=['GET'])
@login_required
def export_classes():
    """API endpoint to export classes data as Excel file"""
    if current_user.role != 'admin':
        return (jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403)
    try:
        from models import SystemSettings
        settings = SystemSettings.query.all()
        settings_dict = {s.key: s.value for s in settings}
        current_term = settings_dict.get('semester', '1st semester')
        current_school_year = settings_dict.get('school_year', '2025-2026')
        classes = Class.query.all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Classes'
        for row_num, cls in enumerate(classes, 1):
            instructor_name = 'Unassigned'
            if cls.instructor:
                if cls.instructor.first_name and cls.instructor.last_name:
                    instructor_name = f'{cls.instructor.first_name} {cls.instructor.last_name}'
                elif cls.instructor.first_name:
                    instructor_name = cls.instructor.first_name
                elif cls.instructor.last_name:
                    instructor_name = cls.instructor.last_name
                else:
                    instructor_name = cls.instructor.username
            course_code = cls.course.code if cls.course else 'Unknown'
            course_description = cls.course.description if cls.course else 'Unknown'
            ws.cell(row=row_num, column=1, value=cls.class_code)
            ws.cell(row=row_num, column=2, value=course_code)
            ws.cell(row=row_num, column=3, value=cls.description or '')
            ws.cell(row=row_num, column=4, value=cls.room_number or '')
            ws.cell(row=row_num, column=5, value=cls.schedule or '')
            ws.cell(row=row_num, column=6, value=instructor_name)
            ws.cell(row=row_num, column=7, value=current_term)
            ws.cell(row=row_num, column=8, value=current_school_year)
        column_widths = [20, 20, 30, 20, 30, 30, 30, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        from flask import send_file
        return send_file(excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='classes_export.xlsx')
    except Exception as e:
        return (jsonify({'success': False, 'message': f'Export failed: {str(e)}'}), 500)

@classes_bp.route('/api/import-classes', methods=['POST'])
@login_required
def import_classes():
    """API endpoint to import classes from XLSX file"""
    if current_user.role != 'admin':
        return (jsonify({'success': False, 'message': 'You do not have permission to perform this action.'}), 403)
    try:
        if 'file' not in request.files:
            return (jsonify({'success': False, 'message': 'No file provided'}), 400)
        file = request.files['file']
        if file.filename == '':
            return (jsonify({'success': False, 'message': 'No file selected'}), 400)
        if not file.filename.lower().endswith('.xlsx'):
            return (jsonify({'success': False, 'message': 'Please upload an XLSX file'}), 400)
        wb = load_workbook(file)
        ws = wb.active
        first_row = [cell.value for cell in ws[1]]
        normalized_headers = [str(value).strip().lower() if value is not None else '' for value in first_row]
        required_headers = ['class code', 'course code', 'room number', 'schedule']
        default_headers = ['class code', 'course code', 'description', 'room number', 'schedule', 'instructor name', 'term', 'school year']
        has_header_row = all((req in normalized_headers for req in required_headers))
        if has_header_row:
            headers = normalized_headers
            data_start_row = 2
        else:
            headers = default_headers.copy()
            if len(headers) < len(first_row):
                headers.extend([''] * (len(first_row) - len(headers)))
            else:
                headers = headers[:len(first_row)]
            data_start_row = 1
        settings = SystemSettings.query.all()
        settings_dict = {s.key: s.value for s in settings}
        default_term = settings_dict.get('semester', '1st semester')
        default_school_year = settings_dict.get('school_year', '2025-2026')
        imported_count = 0
        updated_count = 0
        course_updated_count = 0
        errors = []
        for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), data_start_row):
            try:
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = str(value).strip() if value is not None else ''
                class_code = row_data.get('class code', '').strip()
                course_code = row_data.get('course code', '').strip()
                class_description = row_data.get('description', '').strip()
                room_number = row_data.get('room number', '').strip()
                schedule = row_data.get('schedule', '').strip()
                term = row_data.get('term', '').strip()
                school_year = row_data.get('school year', '').strip()
                if not all([class_code, course_code, room_number, schedule]):
                    errors.append(f'Row {row_num}: Missing required fields for class {class_code}')
                    continue
                course = Course.query.filter_by(code=course_code).first()
                if not course:
                    errors.append(f'Row {row_num}: Course {course_code} not found for class {class_code}')
                    continue
                course_description = row_data.get('course description', '').strip()
                if course_description and course_description != 'No description':
                    course.description = course_description
                    db.session.add(course)
                    course_updated_count += 1
                instructor = None
                instructor_name = row_data.get('instructor name', '').strip()
                if instructor_name and instructor_name != 'Unassigned':
                    instructors = User.query.filter_by(role='instructor').all()
                    for inst in instructors:
                        inst_full_name = ''
                        if inst.first_name and inst.last_name:
                            inst_full_name = f'{inst.first_name} {inst.last_name}'
                        elif inst.first_name:
                            inst_full_name = inst.first_name
                        elif inst.last_name:
                            inst_full_name = inst.last_name
                        else:
                            inst_full_name = inst.username
                        if inst_full_name == instructor_name:
                            instructor = inst
                            break
                    if not instructor:
                        instructor = User.query.filter_by(username=instructor_name, role='instructor').first()
                    if not instructor:
                        errors.append(f'Row {row_num}: Instructor {instructor_name} not found for class {class_code}')
                        continue
                is_valid, error_message = validate_schedule_format(schedule)
                if not is_valid:
                    errors.append(f'Invalid schedule format for class {class_code}: {error_message}')
                    continue
                standardized_schedule = standardize_schedule_days(schedule)
                existing_class = Class.query.filter_by(class_code=class_code).first()
                if existing_class:
                    existing_class.course_id = course.id
                    if class_description:
                        existing_class.description = class_description
                    existing_class.room_number = room_number
                    existing_class.schedule = standardized_schedule
                    existing_class.instructor_id = instructor.id if instructor else None
                    if term:
                        existing_class.term = term
                    if school_year:
                        existing_class.school_year = school_year
                    updated_count += 1
                else:
                    existing_classes = Class.query.filter(Class.class_code != class_code).all()
                    conflict, message = check_schedule_conflict(room_number, standardized_schedule, existing_classes)
                    if conflict:
                        errors.append(f'Row {row_num}: Schedule conflict for class {class_code}: {message}')
                        continue
                    new_class = Class(class_code=class_code, course_id=course.id, description=class_description if class_description else None, room_number=room_number, schedule=standardized_schedule, instructor_id=instructor.id if instructor else None, term=term.lower() if term else default_term, school_year=school_year if school_year else default_school_year, created_at=pst_now_naive())
                    db.session.add(new_class)
                    imported_count += 1
            except Exception as e:
                errors.append(f'Row {row_num}: Error processing class {class_code}: {str(e)}')
                continue
        if imported_count > 0 or updated_count > 0 or course_updated_count > 0:
            db.session.commit()
        message = f'Import completed: {imported_count} new classes added, {updated_count} classes updated, {course_updated_count} course descriptions updated'
        if errors:
            message += f'. {len(errors)} errors occurred.'
        return jsonify({'success': True, 'message': message, 'imported': imported_count, 'updated': updated_count, 'errors': errors})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': f'Import failed: {str(e)}'}), 500)
