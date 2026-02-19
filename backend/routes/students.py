from flask import Blueprint, render_template, redirect, url_for, request, jsonify, flash, current_app, send_file, make_response
from datetime import datetime, date, timedelta
from utils.timezone import get_pst_now, pst_now_naive
import calendar
import json
import os
import uuid
import csv
import io
from io import BytesIO
from werkzeug.utils import secure_filename
import re
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from extensions import db
from models import User, Class, Student, Enrollment, AttendanceRecord, InstructorAttendance, AttendanceLog, FaceEncoding
from forms import StudentForm, EnrollmentForm
from decorators import admin_required
from exceptions import AttendanceValidationError
students_bp = Blueprint('students', __name__, url_prefix='/students')
ALLOWED_DEPARTMENTS = {'BSIT'}

def sanitize_name_for_folder(name):
    """
    Sanitize a name to be safe for use as a folder name.
    Removes special characters and replaces spaces with underscores.
    """
    if not name:
        return 'unknown'
    sanitized = re.sub('[^a-zA-Z0-9\\s_-]', '', name)
    sanitized = re.sub('\\s+', '_', sanitized.strip())
    if not sanitized:
        return 'unknown'
    return sanitized.lower()

@students_bp.route('/enroll', methods=['GET'])
@admin_required
def enroll():
    form = StudentForm()
    return render_template('admin/students.html', form=form)

@students_bp.route('/api/list', methods=['GET'])
def get_students():
    try:
        students = Student.query.options(db.joinedload(Student.enrollments), db.joinedload(Student.face_encodings)).order_by(Student.last_name, Student.first_name).all()
        student_list = []
        for student in students:
            face_encoding = student.face_encodings[0] if student.face_encodings else None
            profile_image = face_encoding.image_path if face_encoding and face_encoding.image_path else None
            student_data = {'id': student.id, 'firstName': student.first_name, 'lastName': student.last_name, 'yearLevel': student.year_level, 'department': student.department, 'enrolledClasses': [e.class_id for e in student.enrollments], 'profileImage': profile_image, 'hasFaceImages': len(student.face_encodings) > 0}
            if hasattr(student, 'middle_name'):
                student_data['middleName'] = student.middle_name
            student_list.append(student_data)
        return jsonify({'success': True, 'students': student_list})
    except Exception as e:
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/<string:student_id>', methods=['GET'])
def get_student(student_id):
    try:
        student = Student.query.get_or_404(student_id)
        face_encoding = student.face_encodings[0] if student.face_encodings else None
        profile_image = face_encoding.image_path if face_encoding and face_encoding.image_path else None
        student_data = {'id': student.id, 'firstName': student.first_name, 'lastName': student.last_name, 'yearLevel': student.year_level, 'department': student.department, 'enrolledClasses': [e.class_id for e in student.enrollments], 'profileImage': profile_image}
        if hasattr(student, 'middle_name'):
            student_data['middleName'] = student.middle_name
        return jsonify({'success': True, 'student': student_data})
    except Exception as e:
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/<string:student_id>', methods=['PUT'])
def update_student(student_id):
    try:
        student = Student.query.get_or_404(student_id)
        data = request.get_json()
        required_fields = ['firstName', 'lastName', 'yearLevel']
        if not data:
            return (jsonify({'success': False, 'message': 'No input data provided'}), 400)
        for field in required_fields:
            if field not in data or not data[field]:
                return (jsonify({'success': False, 'message': f'Missing or empty required field: {field}'}), 400)
        allowed_year_levels = ['1st Year', '2nd Year', '3rd Year', '4th Year']
        if data['yearLevel'] not in allowed_year_levels:
            return (jsonify({'success': False, 'message': 'Invalid Year Level provided.'}), 400)
        department = data.get('department', 'BSIT')
        if department not in ALLOWED_DEPARTMENTS:
            return (jsonify({'success': False, 'message': 'Invalid Department provided.'}), 400)
        student.first_name = data['firstName']
        student.last_name = data['lastName']
        student.year_level = data['yearLevel']
        student.department = department
        if hasattr(student, 'middle_name') and 'middleName' in data:
            student.middle_name = data.get('middleName') or None
        db.session.commit()
        return jsonify({'success': True, 'message': 'Student updated successfully'})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/<string:student_id>', methods=['DELETE'])
def delete_student(student_id):
    try:
        student = Student.query.get_or_404(student_id)
        if student.enrollments:
            return (jsonify({'success': False, 'message': 'Cannot delete student who is enrolled in classes'}), 400)
        db.session.delete(student)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Student deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/create', methods=['POST'])
def create_student():
    """Create a new student"""
    try:
        data = request.get_json()
        required_fields = ['firstName', 'lastName', 'id', 'yearLevel']
        if not data:
            return (jsonify({'success': False, 'message': 'No input data provided'}), 400)
        for field in required_fields:
            if field not in data or not data[field]:
                return (jsonify({'success': False, 'message': f'Missing or empty required field: {field}'}), 400)
        if not re.match('^\\d{2}-\\d{5}$', data['id']):
            return (jsonify({'success': False, 'message': 'Invalid Student ID format. Use YY-XXXXX.'}), 400)
        allowed_year_levels = ['1st Year', '2nd Year', '3rd Year', '4th Year']
        if data['yearLevel'] not in allowed_year_levels:
            return (jsonify({'success': False, 'message': 'Invalid Year Level provided.'}), 400)
        department = data.get('department', 'BSIT')
        if department not in ALLOWED_DEPARTMENTS:
            return (jsonify({'success': False, 'message': 'Invalid Department provided.'}), 400)
        existing_student = Student.query.get(data['id'])
        if existing_student:
            return (jsonify({'success': False, 'message': 'Student ID already exists'}), 400)
        student_kwargs = {'id': data['id'], 'first_name': data['firstName'], 'last_name': data['lastName'], 'year_level': data['yearLevel'], 'department': department}
        if hasattr(Student, 'middle_name') and 'middleName' in data:
            student_kwargs['middle_name'] = data.get('middleName') or None
        student = Student(**student_kwargs)
        try:
            db.session.add(student)
            db.session.commit()
            return jsonify({'success': True, 'message': 'Student created successfully', 'student': {'id': student.id, 'firstName': student.first_name, 'lastName': student.last_name, 'yearLevel': student.year_level, 'department': student.department, 'enrolledClasses': [], 'profileImage': None, **({'middleName': student.middle_name} if hasattr(student, 'middle_name') else {})}})
        except IntegrityError as db_error:
            db.session.rollback()
            return (jsonify({'success': False, 'message': 'Database integrity error: ' + str(db_error)}), 500)
        except Exception as db_error:
            db.session.rollback()
            return (jsonify({'success': False, 'message': 'An unexpected database error occurred: ' + str(db_error)}), 500)
    except Exception as e:
        if db.session.in_transaction():
            db.session.rollback()
        return (jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500)

@students_bp.route('/api/upload-images', methods=['POST'])
def upload_student_images():
    """Upload multiple student images"""
    try:
        student_id = request.form.get('student_id')
        if not student_id:
            return (jsonify({'success': False, 'message': 'Student ID is required'}), 400)
        student = Student.query.get(student_id)
        if not student:
            return (jsonify({'success': False, 'message': 'Student not found'}), 404)
        student_name = f'{student.first_name}_{student.last_name}'
        sanitized_student_name = sanitize_name_for_folder(student_name)
        files = request.files.getlist('images')
        if not files:
            return (jsonify({'success': False, 'message': 'No files provided'}), 400)
        MAX_IMAGES_PER_STUDENT = 20
        face_encodings = FaceEncoding.query.filter_by(student_id=student_id).all()
        if len(face_encodings) + len(files) > MAX_IMAGES_PER_STUDENT:
            return (jsonify({'success': False, 'message': f'Maximum of {MAX_IMAGES_PER_STUDENT} images allowed per student. You currently have {len(face_encodings)} images and are trying to upload {len(files)} more.'}), 400)
        allowed_extensions = {'png', 'jpg', 'jpeg'}
        uploaded_images = []
        errors = []
        for file in files:
            if not file.filename or '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                error_msg = f'File {file.filename} type not allowed. Please upload PNG, JPG, or JPEG'
                errors.append(error_msg)
                continue
            try:
                filename = secure_filename(f'{uuid.uuid4()}_{file.filename}')
                uploads_dir = os.path.join(current_app.static_folder, 'uploads', 'students', sanitized_student_name)
                os.makedirs(uploads_dir, exist_ok=True)
                file_path = os.path.join(uploads_dir, filename)
                file.save(file_path)
                relative_image_path = os.path.join('uploads', 'students', sanitized_student_name, filename).replace('\\', '/')
                face_encoding = FaceEncoding(student_id=student_id, encoding_data=bytes([0] * 128), image_path=relative_image_path, created_at=pst_now_naive())
                db.session.add(face_encoding)
                uploaded_images.append({'id': face_encoding.id, 'filename': filename, 'path': url_for('static', filename=face_encoding.image_path)})
            except Exception as e:
                error_msg = f'Error uploading {file.filename}: {str(e)}'
                errors.append(error_msg)
                continue
        try:
            if uploaded_images:
                db.session.commit()
                return jsonify({'success': True, 'message': f'Successfully uploaded {len(uploaded_images)} image(s).' + (f' Failed to upload {len(errors)} image(s).' if errors else ''), 'images': uploaded_images, 'errors': errors if errors else []})
            else:
                return (jsonify({'success': False, 'message': 'No images were uploaded successfully.', 'errors': errors}), 400)
        except Exception as e:
            db.session.rollback()
            return (jsonify({'success': False, 'message': str(e)}), 500)
    except Exception as e:
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/upload-image', methods=['POST'])
def upload_student_image():
    """Upload a single student image"""
    try:
        if 'image' not in request.files:
            return (jsonify({'success': False, 'message': 'No image file provided'}), 400)
        if 'student_id' not in request.form:
            return (jsonify({'success': False, 'message': 'Student ID is required'}), 400)
        student_id = request.form['student_id']
        file = request.files['image']
        student = Student.query.get(student_id)
        if not student:
            return (jsonify({'success': False, 'message': 'Student not found'}), 404)
        student_name = f'{student.first_name}_{student.last_name}'
        sanitized_student_name = sanitize_name_for_folder(student_name)
        allowed_extensions = {'png', 'jpg', 'jpeg'}
        if not file.filename:
            return (jsonify({'success': False, 'message': 'No file selected'}), 400)
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_extensions:
            return (jsonify({'success': False, 'message': 'File type not allowed. Please upload PNG, JPG, or JPEG'}), 400)
        try:
            filename = secure_filename(f'{uuid.uuid4()}_{file.filename}')
            uploads_dir = os.path.join(current_app.static_folder, 'uploads', 'students', sanitized_student_name)
            os.makedirs(uploads_dir, exist_ok=True)
            file_path = os.path.join(uploads_dir, filename)
            file.save(file_path)
            relative_image_path = os.path.join('uploads', 'students', sanitized_student_name, filename).replace('\\', '/')
            placeholder_encoding = bytes([0] * 128)
            face_encoding = FaceEncoding(student_id=student_id, encoding_data=placeholder_encoding, image_path=relative_image_path, created_at=pst_now_naive())
            try:
                db.session.add(face_encoding)
                face_encoding.encoding_data = placeholder_encoding
                db.session.commit()
                return jsonify({'success': True, 'message': 'Image uploaded successfully', 'image': {'id': face_encoding.id, 'filename': filename, 'path': url_for('static', filename=relative_image_path)}})
            except Exception as db_error:
                db.session.rollback()
                if os.path.exists(file_path):
                    os.remove(file_path)
                return (jsonify({'success': False, 'message': f'Database error: {str(db_error)}'}), 500)
        except Exception as e:
            return (jsonify({'success': False, 'message': str(e)}), 500)
    except Exception as e:
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/images/<string:student_id>', methods=['GET'])
def get_student_images(student_id):
    """Get all facial recognition images for a student"""
    student = Student.query.get(student_id)
    if not student:
        return (jsonify({'success': False, 'message': 'Student not found'}), 404)
    face_encodings = FaceEncoding.query.filter_by(student_id=student_id).all()
    images = []
    for encoding in face_encodings:
        if encoding.image_path:
            image_path_for_url = encoding.image_path.replace('\\', '/')
            if not image_path_for_url.startswith('uploads/students/'):
                image_path_for_url = os.path.join('uploads', 'students', os.path.basename(image_path_for_url)).replace('\\', '/')
            images.append({'id': encoding.id, 'filename': os.path.basename(encoding.image_path), 'path': url_for('static', filename=image_path_for_url), 'createdAt': encoding.created_at.isoformat() if encoding.created_at else None})
    return jsonify({'success': True, 'student': {'id': student.id, 'name': f'{student.first_name} {student.last_name}'}, 'images': images})

@students_bp.route('/api/delete-image/<int:image_id>', methods=['DELETE'])
def delete_student_image(image_id):
    """Deletes a student image by image ID"""
    face_encoding = FaceEncoding.query.get(image_id)
    if not face_encoding:
        return (jsonify({'success': False, 'message': 'Image not found'}), 404)
    try:
        if face_encoding.image_path:
            file_path = os.path.join(current_app.static_folder, face_encoding.image_path)
            if os.path.exists(file_path):
                os.remove(file_path)
        db.session.delete(face_encoding)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Image deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/face-encodings/<string:student_id>', methods=['PUT'])
def save_face_encodings(student_id):
    """Update face encoding data for a student."""
    student = Student.query.get(student_id)
    if not student:
        return (jsonify({'success': False, 'message': 'Student not found'}), 404)
    data = request.get_json()
    if not data or 'encodings' not in data:
        return (jsonify({'success': False, 'message': 'No encoding data provided'}), 400)
    try:
        FaceEncoding.query.filter_by(student_id=student_id).delete()
        for encoding_data in data['encodings']:
            encoding = FaceEncoding(student_id=student_id, encoding_data=encoding_data.encode('utf-8'), created_at=pst_now_naive())
            db.session.add(encoding)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Face encodings saved successfully'})
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/generate-id', methods=['GET'])
def generate_student_id():
    year = get_pst_now().year % 100
    current_year_pattern = f'{year}-'
    max_id = 0
    students = Student.query.filter(Student.id.like(f'{current_year_pattern}%')).all()
    for student in students:
        try:
            id_number = int(student.id.split('-')[1])
            if id_number > max_id:
                max_id = id_number
        except (IndexError, ValueError):
            pass
    next_id = f'{year}-{max_id + 1:05d}'
    return jsonify({'success': True, 'id': next_id})

@students_bp.route('/api/export', methods=['GET'])
def export_students():
    """Export all students to XLSX file with auto-fit column widths"""
    try:
        students = Student.query.order_by(Student.last_name, Student.first_name).all()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Students'
        for row_num, student in enumerate(students, start=1):
            middle_part = ''
            if hasattr(student, 'middle_name') and student.middle_name:
                middle_part = ' ' + str(student.middle_name).strip()
            full_name = f'{student.last_name}, {student.first_name}{middle_part}'
            year_number = re.search('\\d+', student.year_level)
            year = year_number.group() if year_number else ''
            department_year = f'{student.department}-{year}'
            ws.cell(row=row_num, column=1, value=row_num)
            ws.cell(row=row_num, column=2, value=student.id)
            ws.cell(row=row_num, column=3, value=full_name)
            ws.cell(row=row_num, column=4, value=department_year)
        for col_num in range(1, 5):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        filename = f"students_export_{get_pst_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(excel_buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        return (jsonify({'success': False, 'message': str(e)}), 500)

@students_bp.route('/api/import', methods=['POST'])
def import_students():
    """Import students from CSV or XLSX (including XLSX produced by the app export).

    Supports two formats:
    - CSV with headers: 'Student ID', 'First Name', 'Last Name', 'Year Level'
    - XLSX exported by `export_students()` which writes rows: (No header) Number, Student ID, Full Name ("Last, First Middlename"), Department-Year (e.g., "BSIT-4").
    """
    import csv, io
    from openpyxl import load_workbook
    update_existing = request.args.get('update_existing', 'false').lower() == 'true'
    try:
        if 'file' not in request.files:
            return (jsonify({'success': False, 'message': 'No file uploaded.'}), 400)
        file = request.files['file']
        if file.filename == '':
            return (jsonify({'success': False, 'message': 'No file selected.'}), 400)
        filename_lower = file.filename.lower()
        imported_count = 0
        errors = []
        total_rows = 0

        def normalize_year_from_number(num_str):
            mapping = {'1': '1st Year', '2': '2nd Year', '3': '3rd Year', '4': '4th Year'}
            return mapping.get(num_str, num_str + 'th Year' if num_str else '')

        def parse_export_full_name(full_name):
            if not full_name:
                return ('', '', None)
            if ',' in full_name:
                parts = [p.strip() for p in full_name.split(',', 1)]
                last = parts[0]
                rest = parts[1]
                tokens = rest.split()
                first = tokens[0] if tokens else ''
                middle = ' '.join(tokens[1:]) if len(tokens) > 1 else None
                return (first, last, middle)
            tokens = full_name.split()
            if len(tokens) == 1:
                return (tokens[0], '', None)
            return (tokens[0], tokens[-1], ' '.join(tokens[1:-1]) if len(tokens) > 2 else None)
        if filename_lower.endswith('.csv'):
            content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))
            required_columns = ['Student ID', 'First Name', 'Last Name', 'Year Level']
            if not all((col in csv_reader.fieldnames for col in required_columns)):
                return (jsonify({'success': False, 'message': 'Missing required columns.'}), 400)
            for row_num, row in enumerate(csv_reader, start=2):
                total_rows += 1
                student_id = (row.get('Student ID') or '').strip()
                if not student_id:
                    errors.append(f'Row {row_num}: Missing Student ID')
                    continue
                first_name = (row.get('First Name') or '').strip()
                last_name = (row.get('Last Name') or '').strip()
                year_level = (row.get('Year Level') or '').strip()
                student = Student.query.get(student_id)
                if student:
                    if update_existing:
                        student.first_name = first_name or student.first_name
                        student.last_name = last_name or student.last_name
                        student.year_level = year_level or student.year_level
                        imported_count += 1
                    else:
                        errors.append(f'Row {row_num}: Student ID {student_id} already exists')
                else:
                    new_student = Student(id=student_id, first_name=first_name, last_name=last_name, year_level=year_level)
                    db.session.add(new_student)
                    imported_count += 1
        elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(row):
                    continue
                total_rows += 1
                student_id = str(row[1]).strip() if row[1] is not None else ''
                full_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                dept_year = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
                if not student_id:
                    errors.append(f'Row {idx}: Missing Student ID')
                    continue
                first_name, last_name, middle_name = parse_export_full_name(full_name)
                year_number_match = re.search('-(\\d+)$', dept_year)
                year_label = ''
                if year_number_match:
                    year_label = normalize_year_from_number(year_number_match.group(1))
                student = Student.query.get(student_id)
                if student:
                    if update_existing:
                        student.first_name = first_name or student.first_name
                        student.last_name = last_name or student.last_name
                        if middle_name is not None and hasattr(student, 'middle_name'):
                            student.middle_name = middle_name or None
                        student.year_level = year_label or student.year_level
                        imported_count += 1
                    else:
                        errors.append(f'Row {idx}: Student ID {student_id} already exists')
                else:
                    kwargs = {'id': student_id, 'first_name': first_name, 'last_name': last_name, 'year_level': year_label}
                    if hasattr(Student, 'middle_name'):
                        kwargs['middle_name'] = middle_name or None
                    new_student = Student(**kwargs)
                    db.session.add(new_student)
                    imported_count += 1
        else:
            return (jsonify({'success': False, 'message': 'Invalid file type. Accepted: .csv, .xlsx, .xls'}), 400)
        if imported_count > 0:
            db.session.commit()
        response_data = {'success': True, 'imported_count': imported_count, 'total_rows': total_rows, 'errors': errors}
        return jsonify(response_data)
    except Exception as e:
        db.session.rollback()
        return (jsonify({'success': False, 'message': str(e)}), 500)
